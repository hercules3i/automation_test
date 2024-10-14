import openpyxl
from Task import TaskSimpleData
from Bom import Bom, Warehouse
import os
work_book = openpyxl.load_workbook(os.path.join("data","Data Test BOM.xlsx"))

list_workings = ["Công đoạn kéo", "Công đoạn dập và ren",
                 "Nhiệt", "Xi mạ", "Đóng gói"]


def LoadSimpleData() -> list:
    sheet = work_book["Không có thuộc tính mở rộng"]
    max_column = sheet.max_column
    max_row = sheet.max_row

    list_tasks = list()

    latest_machines = None
    latest_workers = None
    latest_work_order = ''
    latest_shift = None
    latest_in_or_out = None
    for row in range(4, max_row):
        #
        working = sheet.cell(row, 1).value.__str__().strip()
        if working in list_workings:
            task = TaskSimpleData()
            task.name_of_task = working
            list_tasks.append(task)
            latest_machines = None
            latest_workers = None
            latest_work_order = ''
            latest_shift = None
            latest_in_or_out = None
            continue

        #collect bom-details
        work_order = sheet.cell(row, 1).value #"none"
        if work_order is None and latest_work_order != '': work_order = latest_work_order

        shift = sheet.cell(row, 2).value #ca2 13:30 - 17:30 
        if shift is None: shift = latest_shift

        machines = sheet.cell(row, 3).value #BF24-01
        if machines is None: machines = latest_machines

        workers = sheet.cell(row, 4).value  # Phamba vuong
        if workers is None: workers = latest_workers

        in_or_out = sheet.cell(row, 5).value #nhap
        if in_or_out is None: in_or_out = latest_in_or_out

        #create new_bom with new order
        if (work_order != latest_work_order and latest_work_order != '' and work_order is not None
                # create new_bom with the newest work_order
                or latest_work_order == ''):
            new_bom = Bom()
            task.list_boms.append(new_bom)

        #create new_bom with new shift
        if (work_order == latest_work_order and latest_work_order != ''
                and shift != latest_shift and latest_shift is not None and shift is not None):
            new_bom = Bom()
            new_bom.work_order = latest_work_order
            task.list_boms.append(new_bom)

        #create new_bom with new workers
        if (work_order == latest_work_order and latest_work_order != ''
                and shift == latest_shift and latest_shift is not None
                and workers != latest_workers and latest_workers is not None):
            new_bom = Bom()
            new_bom.work_order = latest_work_order
            task.list_boms.append(new_bom)

        #create new case with new in_or_out
        if (work_order == latest_work_order and latest_work_order != ''
                and shift == latest_shift and latest_shift is not None
                and workers == latest_workers and latest_workers is not None
                and in_or_out != latest_in_or_out and latest_in_or_out is not None):
            new_bom = Bom()
            new_bom.work_order = latest_work_order
            task.list_boms.append(new_bom)


        #add warehouse delails
        #work_order
        if latest_work_order == '' or work_order is not None:
            if work_order is not None:
                new_work_order = ''
                for word in work_order.split(' '):
                    if word == '':
                        continue
                    word = word.strip()
                    new_work_order += word + ' '

                work_order = new_work_order.strip(' ')

            new_bom.work_order = work_order
            latest_work_order = work_order

        #shift
        if shift is not None:
            new_bom.shift = shift
            latest_shift = shift

        #machines
        if machines is not None:
            new_bom.machines = machines.__str__().split(",")
            latest_machines = machines
        else:
            new_bom.machines = latest_machines

        #workers
        if workers is not None:
            new_bom.workers = workers.__str__().split(",")
            latest_workers = workers

        #in_or_out
        if in_or_out is not None:
            new_bom.in_or_out = in_or_out
            latest_in_or_out = in_or_out

        #define warehouse obj
        if in_or_out == "Nhập":
            warehouse_obj = new_bom.warehouse_in
        elif in_or_out == "Xuất":
            warehouse_obj = new_bom.warehouse_out

        #detect warehouse's input and output
        in_out_value = sheet.cell(row, 6).value
        if in_out_value.__str__().strip() == "Đầu vào":
            warehouse_prod = warehouse_obj.input
        elif in_out_value.__str__().strip() == "Đầu ra":
            warehouse_prod = warehouse_obj.output

        #add product details
        product = warehouse_obj.Product()
        product.name = sheet.cell(row, 7).value
        if product.name is not None:
            product.name = product.name.rstrip("\n\t")
        product.quantity = sheet.cell(row, 8).value
        product.unit = sheet.cell(row, 9).value
        product.specification = sheet.cell(row, 10).value
        product.tail = sheet.cell(row, 11).value
        product.not_good = sheet.cell(row, 12).value

        if not product.IsNone():
            warehouse_prod.product_list.append(product)

    return list_tasks


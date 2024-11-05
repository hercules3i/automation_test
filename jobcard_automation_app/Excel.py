import openpyxl
from JobcardClass import TabJob,CardJob,Staff,Task
import os
from father import JobCardSimpleData
work_book = openpyxl.load_workbook(os.path.join("data","data thẻ việc (2).xlsx"))
# print(work_book)
# work_book = openpyxl.load_workbook(os.path.join("Downloads","automation_test-main (1)","data","data thẻ việc.xlsx"))
list_sheets = ["Bảng việc ","Thêm thẻ việc","Thêm nhân sự ","Thêm đầu mục"]
def LoadSimpleData():
    try:
        sheet_tab_job = work_book[list_sheets[0]]
        sheet_card_job = work_book[list_sheets[1]]
        sheet_staff = work_book[list_sheets[2]]
        sheet_task = work_book[list_sheets[3]]
        jobcard = JobCardSimpleData()
        list_tabjobs = create_tabjob_object(sheet_tab_job,jobcard)
        list_jobcards = create_cardjob_object(sheet_card_job,jobcard)
        list_staffs = create_staff_object(sheet_staff,jobcard)
        list_tasks = create_task_object(sheet_task,jobcard)
        result = {
            "list_tabjobs":list_tabjobs,
            "list_jobcards":list_jobcards,
            "list_staffs":list_staffs,
            "list_tasks":list_tasks
        }
        return result
    except Exception:
        return "Error"
        # logging.error(f"Lỗi khi thực hiện {task_name}: {e}")

    
def create_tabjob_object(sheet,jobcard):
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        tab_name = sheet.cell(row, 2).value
        branch = sheet.cell(row,3).value
        department = sheet.cell(row,4).value
        tab_type = sheet.cell(row,5).value
        start_date = sheet.cell(row,6).value
        end_date = sheet.cell(row,7).value
        new_tab_job = TabJob(tab_name = tab_name, branch=branch, department=department,tab_type=tab_type,start_date=start_date,end_date=end_date)
        jobcard.list_tabjobs.append(new_tab_job)
    return jobcard.list_tabjobs

def create_cardjob_object(sheet,jobcard):
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        card_name = sheet.cell(row, 2).value
        print(card_name)
        start_date = sheet.cell(row,3).value
        end_date = sheet.cell(row,4).value
        card_status = sheet.cell(row,5).value
        weight = sheet.cell(row,6).value
        tab_job = sheet.cell(row,7).value
        category_job = sheet.cell(row,8).value
        priority = sheet.cell(row,9).value
        job_type = sheet.cell(row,10).value
        funding = sheet.cell(row,11).value
        money_type = sheet.cell(row,12).value
        cycle = sheet.cell(row,13).value
        workflow = sheet.cell(row,14).value
        new_card_job = CardJob(card_name = card_name, start_date=start_date,end_date = end_date, card_status=card_status,weight=weight,tab_job=tab_job,category_job=category_job,priority=priority,job_type=job_type,funding=funding,money_type=money_type,cycle=cycle,workflow=workflow)
        
        jobcard.list_jobcards.append(new_card_job)
    return jobcard.list_jobcards

def create_staff_object(sheet, jobcard):
    max_row = sheet.max_row
    old_jobcard_name = ""
    for row in range(3, max_row + 1):
        card_job = sheet.cell(row, 2).value
        if card_job and card_job != None:
            old_jobcard_name = card_job
        branch_agency = sheet.cell(row, 3).value
        department_group = sheet.cell(row, 4).value
        try:
            department_group = department_group.split(",")
        except:
            print("Chi co mot phong ban")
        staff_name = sheet.cell(row, 5).value
        try:
            list_staffs = staff_name.split(",")
            new_staff = Staff(old_jobcard_name, branch_agency, department_group, list_staffs)
            jobcard.list_staffs.append(new_staff)
        except:
            print("Chi co mot nhan vien")
            new_staff = Staff(old_jobcard_name, branch_agency, department_group, list(staff_name))
            jobcard.list_staffs.append(new_staff)
    return jobcard.list_staffs


def create_task_object(sheet, jobcard):
    max_row = sheet.max_row
    max_col = sheet.max_column 
    for row in range(3, max_row + 1): 
        card_job = sheet.cell(row, 2).value  
        j = 3
        for col in range(j, max_col + 1, 5): 
            task_name = sheet.cell(row, col).value
            weight = sheet.cell(row, col + 1).value
            staff = sheet.cell(row, col + 2).value
            duration = sheet.cell(row, col + 3).value
            unit = sheet.cell(row, col + 4).value
            j += 4
            new_task = Task(card_job, task_name, weight, staff,duration,unit)
            jobcard.list_tasks.append(new_task)
    return jobcard.list_tasks

# result = LoadSimpleData()


# for i in range(len(result["list_jobcards"])):
#     object = result["list_jobcards"][i]
#     try:
#         # Access the attributes of the object
#         print(f"{object.name} | {object.start_date} | {object.end_date} | {object.card_status} | {object.tab_job} | {object.category_job}")
#     except Exception as e:  # Catch the exception and store it in variable 'e'
#         print(e)  # Print the exception message
